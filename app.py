from flask import Flask, request, redirect, url_for, render_template_string, jsonify, session
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os
import tempfile
import openpyxl
import pdfplumber
import re

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'easystock-secret-key-change-me')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Многокомпанейность: у каждой компании СВОЯ база данных.
# DATABASE_URL — это база-реестр, в ней лежит только список компаний
# и строки подключения к их базам. Товаров и операций в реестре нет.
REGISTRY_URL = os.environ.get('DATABASE_URL', '')
USE_PG = REGISTRY_URL.startswith('postgres')

# Пароли стартовых учёток компании. Задаются при её создании,
# эти значения — только запасные по умолчанию.
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')
SELLER_PASSWORD = os.environ.get('SELLER_PASSWORD', 'seller123')

# Владелец сервиса. Заводит компании, внутрь их данных не заходит.
SUPERADMIN_USERNAME = os.environ.get('SUPERADMIN_USERNAME', 'superadmin')
SUPERADMIN_PASSWORD = os.environ.get('SUPERADMIN_PASSWORD', 'super123')

if USE_PG:
    import psycopg2
    import psycopg2.extras
    if REGISTRY_URL.startswith('postgres://'):
        REGISTRY_URL = REGISTRY_URL.replace('postgres://', 'postgresql://', 1)
    if not os.environ.get('SECRET_KEY'):
        print('[EasyStock] ВНИМАНИЕ: SECRET_KEY не задан, сессии уязвимы.')
    if not os.environ.get('SUPERADMIN_PASSWORD'):
        print('[EasyStock] ВНИМАНИЕ: SUPERADMIN_PASSWORD не задан, используется '
              'super123. Задайте его в переменных окружения.')
else:
    import sqlite3
    REGISTRY_URL = os.path.join(BASE_DIR, 'registry.db')


def connect(target):
    """Подключение к произвольной базе.

    target — строка postgresql://... на проде либо имя файла SQLite локально.
    """
    if USE_PG:
        conn = psycopg2.connect(target)
        conn.autocommit = False
        return conn
    conn = sqlite3.connect(target if os.path.isabs(target) else os.path.join(BASE_DIR, target))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def get_registry():
    """Соединение с базой-реестром компаний."""
    return connect(REGISTRY_URL)


def get_db():
    """Соединение с базой ТЕКУЩЕЙ компании (определяется по сессии).

    Роуты работы с товарами вызывают именно её, поэтому каждый запрос
    автоматически попадает в базу своей компании и никогда — в чужую.
    """
    company = current_company()
    if company is None:
        raise LookupError('Компания не определена: нет активной сессии')
    return connect(company['db_url'])


def db_execute(conn, query, params=None):
    """Execute query compatible with both SQLite and PostgreSQL."""
    if USE_PG:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # Convert ? placeholders to %s for PostgreSQL
        query = query.replace('?', '%s')
        # params=None отключает подстановку в psycopg2. С пустым кортежем
        # она включается, и любой литерал % в SQL (например, LIKE '%текст%')
        # ломает запрос.
        cur.execute(query, params if params else None)
        return cur
    else:
        return conn.execute(query, params or ())


def db_fetchall(conn, query, params=None):
    cur = db_execute(conn, query, params)
    return cur.fetchall()


def db_fetchone(conn, query, params=None):
    cur = db_execute(conn, query, params)
    return cur.fetchone()


def ensure_admin_exists(conn):
    """Гарантирует, что в системе есть хотя бы один администратор.

    На базах, созданных до появления ролей, ALTER TABLE проставляет всем
    пользователям роль 'seller' — включая учётку admin. Повторный INSERT
    её не исправляет (логин уже занят), и система остаётся без админа.
    """
    row = db_fetchone(conn, "SELECT COUNT(*) AS n FROM users WHERE role = 'admin'")
    count = row['n'] if row else 0
    if count == 0:
        db_execute(conn, "UPDATE users SET role = 'admin' WHERE username = 'admin'")
        print("[EasyStock] Учётке 'admin' восстановлена роль администратора.")


# --- Реестр компаний -------------------------------------------------------

# Строки подключения кэшируются в памяти: иначе каждый запрос страницы
# требовал бы лишнего обращения к реестру. В сессию их класть нельзя —
# cookie подписан, но не зашифрован, и пароль от базы утёк бы в браузер.
_company_cache = {}


def _remember(row):
    if row is None:
        return None
    data = dict(row)
    _company_cache[data['id']] = data
    return data


def forget_companies():
    """Сбросить кэш после изменений в реестре."""
    _company_cache.clear()


def find_company_by_code(code):
    conn = get_registry()
    row = db_fetchone(conn,
                      'SELECT * FROM companies WHERE code = ? AND is_active = 1',
                      (code.strip().lower(),))
    conn.close()
    return _remember(row)


def get_company(company_id):
    if company_id in _company_cache:
        return _company_cache[company_id]
    conn = get_registry()
    row = db_fetchone(conn, 'SELECT * FROM companies WHERE id = ?', (company_id,))
    conn.close()
    return _remember(row)


def all_companies():
    conn = get_registry()
    rows = db_fetchall(conn, 'SELECT * FROM companies ORDER BY name')
    conn.close()
    return [dict(r) for r in rows]


def current_company():
    """Компания текущей сессии."""
    company_id = session.get('company_id')
    if not company_id:
        return None
    return get_company(company_id)


def init_registry():
    """Создаёт таблицу компаний. Данных компаний в реестре не хранится."""
    conn = get_registry()
    if USE_PG:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS companies (
                id SERIAL PRIMARY KEY,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                db_url TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT (CURRENT_DATE::text)
            )
        ''')
        conn.commit()
        cur.close()
    else:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                db_url TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT (date('now'))
            );
        ''')
        conn.commit()
        # Локально сразу заводим демо-компанию на старой базе,
        # чтобы разработка не требовала ручной настройки.
        existing = db_fetchone(conn, 'SELECT COUNT(*) AS n FROM companies')
        if (existing['n'] if existing else 0) == 0:
            conn.execute(
                'INSERT INTO companies (code, name, db_url) VALUES (?, ?, ?)',
                ('demo', 'Демо-компания', 'inventory.db'))
            conn.commit()
            init_company_db('inventory.db', ADMIN_PASSWORD, SELLER_PASSWORD)
            print("[EasyStock] Создана локальная компания 'demo' на базе inventory.db")
    conn.close()


def init_company_db(target, admin_password, seller_password):
    """Разворачивает схему в базе компании и заводит две стартовые учётки."""
    conn = connect(target)
    if USE_PG:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id SERIAL PRIMARY KEY,
                product_id INTEGER NOT NULL REFERENCES products(id),
                type TEXT NOT NULL CHECK(type IN ('in', 'out')),
                quantity REAL NOT NULL,
                price REAL NOT NULL,
                date TEXT NOT NULL DEFAULT (CURRENT_DATE::text)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'seller'
            )
        ''')
        conn.commit()
        # Default users
        try:
            cur.execute('INSERT INTO users (username, password, role) VALUES (%s, %s, %s)',
                        ('admin', generate_password_hash(admin_password), 'admin'))
            conn.commit()
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
        try:
            cur.execute('INSERT INTO users (username, password, role) VALUES (%s, %s, %s)',
                        ('seller', generate_password_hash(seller_password), 'seller'))
            conn.commit()
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
        ensure_admin_exists(conn)
        conn.commit()
        cur.close()
    else:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                type TEXT NOT NULL CHECK(type IN ('in', 'out')),
                quantity REAL NOT NULL,
                price REAL NOT NULL,
                date TEXT NOT NULL DEFAULT (date('now')),
                FOREIGN KEY (product_id) REFERENCES products(id)
            );
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'seller'
            );
        ''')
        try:
            conn.execute('ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT "seller"')
        except sqlite3.OperationalError:
            pass  # колонка уже есть
        try:
            conn.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                         ('admin', generate_password_hash(admin_password), 'admin'))
        except sqlite3.IntegrityError:
            pass
        try:
            conn.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                         ('seller', generate_password_hash(seller_password), 'seller'))
        except sqlite3.IntegrityError:
            pass
        ensure_admin_exists(conn)
        conn.commit()
    conn.close()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        # Компания обязательна: без неё непонятно, в какую базу идти,
        # а «просто вошедший» пользователь не должен видеть ничего.
        if 'user_id' not in session or 'company_id' not in session:
            return redirect(url_for('login'))
        if current_company() is None:
            session.clear()
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or 'company_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def superadmin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_superadmin'):
            return redirect(url_for('superadmin_login'))
        return f(*args, **kwargs)
    return decorated

LOGIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EasyStock - Вход</title>
<style>
:root {
  --bg: #0e1117; --bg2: #161b27; --bg3: #1e2535;
  --border: #2a3347; --border2: #374056;
  --text: #e8ecf4; --text2: #8b95b0; --text3: #5a6480;
  --accent: #4f8ef7; --accent2: #3b6fd4;
  --green: #2ecc8a;
  --red: #f74f4f;
  --radius: 10px; --radius2: 6px;
}
* { margin:0; padding:0; box-sizing:border-box; }
body {
  font-family: 'Segoe UI', sans-serif;
  background: var(--bg);
  color: var(--text);
  min-height: 100vh;
  display: flex;
  align-items: center;
  justify-content: center;
}
.login-card {
  background: var(--bg2);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 40px;
  width: 380px;
  max-width: 95vw;
  box-shadow: 0 4px 24px rgba(0,0,0,0.4);
}
.login-logo {
  text-align: center;
  margin-bottom: 30px;
}
.login-logo h1 {
  font-size: 24px;
  color: var(--accent);
  letter-spacing: 0.05em;
}
.login-logo p {
  font-size: 12px;
  color: var(--text3);
  margin-top: 4px;
}
.form-group { margin-bottom: 16px; }
.form-label {
  display: block; font-size: 11px; font-weight: 600;
  color: var(--text2); margin-bottom: 5px;
}
.form-control {
  width: 100%; padding: 10px 14px;
  background: var(--bg3); border: 1px solid var(--border2);
  border-radius: var(--radius2); color: var(--text);
  font-family: inherit; font-size: 14px;
  outline: none; transition: border-color 0.15s;
}
.form-control:focus { border-color: var(--accent); }
.btn-login {
  width: 100%; padding: 12px;
  background: var(--accent); color: #fff;
  border: none; border-radius: var(--radius2);
  font-size: 14px; font-weight: 600;
  cursor: pointer; font-family: inherit;
  transition: background 0.15s;
}
.btn-login:hover { background: var(--accent2); }
.error-msg {
  background: rgba(247,79,79,0.1);
  border: 1px solid rgba(247,79,79,0.3);
  color: var(--red);
  padding: 8px 12px;
  border-radius: var(--radius2);
  font-size: 12px;
  margin-bottom: 16px;
  text-align: center;
}
@media (max-width: 480px) {
  body { padding: 16px; align-items: flex-start; padding-top: 8vh; }
  .login-card { padding: 28px 22px; width: 100%; }
  /* 16px не даёт iOS зумить страницу при фокусе на поле */
  .form-control { font-size: 16px; padding: 12px 14px; }
  .btn-login { padding: 14px; font-size: 15px; }
}
</style>
</head>
<body>
<div class="login-card">
  <div class="login-logo">
    <h1>EasyStock</h1>
    <p>Войдите в систему</p>
  </div>
  {% if error %}
  <div class="error-msg">{{ error }}</div>
  {% endif %}
  <form method="POST">
    <div class="form-group">
      <label class="form-label">Код компании</label>
      <input class="form-control" type="text" name="company" placeholder="Например: acme"
             value="{{ company_code or '' }}" required autofocus autocapitalize="off" autocorrect="off">
    </div>
    <div class="form-group">
      <label class="form-label">Логин</label>
      <input class="form-control" type="text" name="username" placeholder="Введите логин" required>
    </div>
    <div class="form-group">
      <label class="form-label">Пароль</label>
      <input class="form-control" type="password" name="password" placeholder="Введите пароль" required>
    </div>
    <button type="submit" class="btn-login">Войти</button>
  </form>
  <div style="margin-top:18px;text-align:center;">
    <a href="/superadmin/login" style="font-size:11px;color:var(--text3);text-decoration:none;">
      Вход для владельца сервиса
    </a>
  </div>
  {% if is_local %}
  <div style="margin-top:24px;border-top:1px solid var(--border);padding-top:16px;">
    <div style="font-size:11px;color:var(--text3);margin-bottom:10px;text-align:center;">Быстрый вход (только локально):</div>
    <div style="display:flex;gap:8px;">
      <button onclick="fillLogin('admin','admin123')" style="flex:1;padding:10px;background:var(--bg3);border:1px solid var(--border2);border-radius:var(--radius2);color:var(--text);cursor:pointer;font-family:inherit;font-size:12px;transition:all 0.15s;">
        <div style="font-weight:700;color:var(--accent);">Админ</div>
        <div style="color:var(--text3);font-size:10px;margin-top:2px;">admin / admin123</div>
      </button>
      <button onclick="fillLogin('seller','seller123')" style="flex:1;padding:10px;background:var(--bg3);border:1px solid var(--border2);border-radius:var(--radius2);color:var(--text);cursor:pointer;font-family:inherit;font-size:12px;transition:all 0.15s;">
        <div style="font-weight:700;color:var(--green);">Продавец</div>
        <div style="color:var(--text3);font-size:10px;margin-top:2px;">seller / seller123</div>
      </button>
    </div>
  </div>
  {% endif %}
</div>
{% if is_local %}
<script>
function fillLogin(u, p) {
  document.querySelector('input[name="username"]').value = u;
  document.querySelector('input[name="password"]').value = p;
  document.querySelector('form').submit();
}
</script>
{% endif %}
</body>
</html>
'''

TEMPLATE = '''
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EasyStock</title>
<style>
:root {
  --bg: #0e1117;
  --bg2: #161b27;
  --bg3: #1e2535;
  --bg4: #252d3f;
  --border: #2a3347;
  --border2: #374056;
  --text: #e8ecf4;
  --text2: #8b95b0;
  --text3: #5a6480;
  --accent: #4f8ef7;
  --accent2: #3b6fd4;
  --green: #2ecc8a;
  --green2: #1a7a52;
  --red: #f74f4f;
  --red2: #8a1a1a;
  --yellow: #f7c44f;
  --purple: #9b6dff;
  --radius: 10px;
  --radius2: 6px;
  --shadow: 0 4px 24px rgba(0,0,0,0.4);
  --mono: 'Segoe UI', monospace;
}

* { margin:0; padding:0; box-sizing:border-box; }

body {
  font-family: 'Segoe UI', sans-serif;
  background: var(--bg);
  color: var(--text);
  min-height: 100vh;
  min-height: 100dvh;
  display: flex;
  font-size: 14px;
  overflow-x: hidden;
}

/* SIDEBAR */
.sidebar {
  width: 220px;
  min-height: 100vh;
  background: var(--bg2);
  border-right: 1px solid var(--border);
  display: flex;
  flex-direction: column;
  position: fixed;
  top: 0; left: 0; bottom: 0;
  z-index: 100;
}
.logo {
  padding: 20px 18px 16px;
  border-bottom: 1px solid var(--border);
}
.logo-title {
  font-size: 16px;
  font-weight: 700;
  color: var(--accent);
  letter-spacing: 0.05em;
}
.logo-sub {
  font-size: 10px;
  color: var(--text3);
  margin-top: 2px;
}
.nav { padding: 12px 0; flex: 1; overflow-y: auto; }
.nav-item {
  display: flex;
  align-items: center;
  gap: 10px;
  padding: 11px 18px;
  cursor: pointer;
  color: var(--text2);
  font-size: 13px;
  font-weight: 500;
  transition: all 0.15s;
  border-left: 2px solid transparent;
  text-decoration: none;
}
.nav-item:hover { background: var(--bg3); color: var(--text); }
.nav-item.active { background: var(--bg3); color: var(--accent); border-left-color: var(--accent); }
.nav-item .icon { font-size: 15px; width: 18px; text-align: center; }

/* MAIN */
.main {
  margin-left: 220px;
  flex: 1;
  display: flex;
  flex-direction: column;
  min-height: 100vh;
  /* Без min-width:0 flex-элемент не сжимается уже своего содержимого,
     и широкая таблица растягивает всю страницу вбок. */
  min-width: 0;
}
.page > * { min-width: 0; }
.topbar {
  height: 56px;
  background: var(--bg2);
  border-bottom: 1px solid var(--border);
  display: flex;
  align-items: center;
  padding: 0 24px;
  position: sticky; top: 0; z-index: 50;
}
.topbar-title { font-size: 16px; font-weight: 700; }

.page { padding: 24px; display: none; }
.page.active { display: block; }

/* STAT CARDS */
.stats-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 20px; }
.stat-card {
  background: var(--bg2);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 18px;
  position: relative;
  overflow: hidden;
}
.stat-card::before {
  content: '';
  position: absolute; top: 0; left: 0; right: 0; height: 2px;
}
.stat-card.accent::before { background: var(--accent); }
.stat-card.green::before { background: var(--green); }
.stat-card.red::before { background: var(--red); }
.stat-card.yellow::before { background: var(--yellow); }
.stat-label { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: var(--text3); }
.stat-value { font-size: 26px; font-weight: 700; margin-top: 6px; font-family: var(--mono); }
.stat-sub { font-size: 11px; color: var(--text3); margin-top: 4px; }
.stat-icon { position: absolute; right: 16px; top: 16px; font-size: 28px; opacity: 0.15; }

/* CARD */
.card {
  background: var(--bg2);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  overflow: hidden;
  margin-bottom: 20px;
}
.card-header {
  padding: 14px 18px;
  border-bottom: 1px solid var(--border);
  display: flex; align-items: center; justify-content: space-between;
}
.card-title { font-size: 13px; font-weight: 700; }
.card-body { padding: 18px; }

/* TABLE */
.tbl { width: 100%; border-collapse: collapse; }
.tbl th {
  text-align: left; padding: 10px 14px;
  font-size: 10px; font-weight: 700; letter-spacing: 0.08em;
  text-transform: uppercase; color: var(--text3);
  background: var(--bg3);
  border-bottom: 1px solid var(--border);
}
.tbl td {
  padding: 11px 14px; font-size: 12px;
  border-bottom: 1px solid var(--border);
  color: var(--text);
}
.tbl tr:last-child td { border-bottom: none; }
.tbl tr:hover td { background: var(--bg3); }
.tbl .num { text-align: right; font-family: var(--mono); }

/* BADGE */
.badge {
  display: inline-flex; align-items: center;
  padding: 2px 8px; border-radius: 20px;
  font-size: 10px; font-weight: 700;
}
.badge-in { background: rgba(46,204,138,0.15); color: var(--green); }
.badge-out { background: rgba(247,79,79,0.15); color: var(--red); }

/* BUTTONS */
.btn {
  display: inline-flex; align-items: center; gap: 6px;
  padding: 8px 14px; border-radius: var(--radius2);
  font-size: 12px; font-weight: 600; cursor: pointer;
  border: none; font-family: inherit; transition: all 0.15s;
  color: #fff;
}
.btn-primary { background: var(--accent); }
.btn-primary:hover { background: var(--accent2); }
.btn-success { background: var(--green2); color: var(--green); border: 1px solid var(--green2); }
.btn-success:hover { background: var(--green); color: #0e1117; }
.btn-danger { background: var(--red2); color: var(--red); border: 1px solid var(--red2); }
.btn-danger:hover { background: var(--red); color: #fff; }
.btn-secondary { background: var(--bg3); color: var(--text); border: 1px solid var(--border2); }
.btn-secondary:hover { background: var(--bg4); }
.btn-sm { padding: 5px 10px; font-size: 11px; }

/* FORM */
.form-group { margin-bottom: 14px; }
.form-label { display: block; font-size: 11px; font-weight: 600; color: var(--text2); margin-bottom: 5px; }
.form-control {
  width: 100%; padding: 8px 12px;
  background: var(--bg3); border: 1px solid var(--border2);
  border-radius: var(--radius2); color: var(--text);
  font-family: inherit; font-size: 13px;
  transition: border-color 0.15s;
  outline: none;
}
.form-control:focus { border-color: var(--accent); }
.form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
.form-row3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; }

/* MODAL */
.modal-overlay {
  position: fixed; inset: 0; background: rgba(0,0,0,0.7);
  z-index: 1000; display: none;
  align-items: center; justify-content: center;
}
.modal-overlay.open { display: flex; }
.modal {
  background: var(--bg2); border: 1px solid var(--border2);
  border-radius: var(--radius); width: 500px; max-width: 95vw;
  max-height: 90vh; overflow-y: auto;
  box-shadow: var(--shadow);
}
.modal-header {
  padding: 18px 22px; border-bottom: 1px solid var(--border);
  display: flex; align-items: center; justify-content: space-between;
}
.modal-title { font-size: 15px; font-weight: 700; }
.modal-close { cursor: pointer; color: var(--text3); font-size: 18px; }
.modal-close:hover { color: var(--text); }
.modal-body { padding: 22px; }
.modal-footer {
  padding: 14px 22px; border-top: 1px solid var(--border);
  display: flex; justify-content: flex-end; gap: 8px;
}

/* NOTIFICATION */
.notif {
  position: fixed; top: 20px; right: 20px; z-index: 9999;
  background: var(--bg2); border: 1px solid var(--border2);
  border-radius: var(--radius); padding: 12px 18px;
  box-shadow: var(--shadow); font-size: 13px;
  transform: translateX(120%); transition: transform 0.3s;
  display: flex; align-items: center; gap: 10px;
}
.notif.show { transform: translateX(0); }
.notif.success { border-left: 3px solid var(--green); }
.notif.error { border-left: 3px solid var(--red); }

.profit-pos { color: var(--green); font-weight: 600; }
.profit-neg { color: var(--red); font-weight: 600; }

/* PAGE HEAD */
.page-head {
  display: flex; justify-content: space-between;
  align-items: center; gap: 12px; margin-bottom: 20px;
}
.page-head-actions { display: flex; gap: 8px; }

/* BURGER + BACKDROP (мобильное меню) */
.burger {
  display: none;
  background: none; border: none; color: var(--text);
  font-size: 22px; cursor: pointer; padding: 4px 12px 4px 0;
  line-height: 1; font-family: inherit;
}
.sidebar-backdrop {
  display: none;
  position: fixed; inset: 0;
  background: rgba(0,0,0,0.6);
  z-index: 99;
}
.sidebar-backdrop.open { display: block; }

/* SCROLLBAR */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 3px; }

/* ===== ПЛАНШЕТ ===== */
@media (max-width: 1024px) {
  .stats-grid { grid-template-columns: 1fr 1fr; }
  .page { padding: 18px; }
}

/* ===== ТЕЛЕФОН ===== */
@media (max-width: 768px) {
  /* Меню выезжает поверх контента, а не сжимает его */
  .sidebar {
    width: 260px;
    transform: translateX(-100%);
    transition: transform 0.25s ease;
    box-shadow: var(--shadow);
  }
  .sidebar.open { transform: translateX(0); }
  .main { margin-left: 0; }
  .burger { display: block; }

  .topbar { padding: 0 14px; height: 52px; }
  .topbar-title { font-size: 15px; }
  .page { padding: 14px; }

  /* Заголовок страницы и кнопки — в столбик, кнопки на всю ширину */
  .page-head { flex-direction: column; align-items: stretch; gap: 10px; }
  .page-head-actions { width: 100%; }
  .page-head-actions .btn { flex: 1; justify-content: center; }

  /* Таблицы прокручиваются внутри карточки, страница вбок не едет */
  .card { overflow-x: auto; -webkit-overflow-scrolling: touch; }
  .tbl { min-width: 560px; }

  .stat-value { font-size: 21px; }
  .stat-icon { font-size: 22px; right: 12px; top: 12px; }
  .stat-card { padding: 14px; }
  .stats-grid { gap: 10px; }

  /* Формы — в одну колонку */
  .form-row, .form-row3 { grid-template-columns: 1fr; }

  /* 16px не даёт iOS зумить страницу при фокусе на поле */
  .form-control { font-size: 16px; padding: 10px 12px; }

  /* Удобные для пальца кнопки */
  .btn { padding: 10px 14px; min-height: 40px; }
  .btn-sm { padding: 8px 10px; min-height: 34px; }

  /* Модалки — на всю ширину экрана снизу */
  .modal-overlay { align-items: flex-end; }
  .modal {
    width: 100% !important; max-width: 100%;
    max-height: 88vh;
    border-radius: var(--radius) var(--radius) 0 0;
  }
  .modal-header { padding: 14px 16px; }
  .modal-body { padding: 16px; }
  .modal-footer { padding: 12px 16px; }
  .modal-footer .btn { flex: 1; justify-content: center; }

  .notif { left: 12px; right: 12px; top: 12px; }
  .notif.show { transform: translateY(0); }
  .notif:not(.show) { transform: translateY(-150%); }
}

/* ===== УЗКИЙ ТЕЛЕФОН ===== */
@media (max-width: 400px) {
  .stats-grid { grid-template-columns: 1fr; }
  .page { padding: 12px; }
}
</style>
</head>
<body>

<!-- SIDEBAR -->
<div class="sidebar-backdrop" id="sidebarBackdrop" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="logo">
    <div class="logo-title">EasyStock</div>
    <div class="logo-sub" title="{{ company_name }}">{{ company_name or 'Учёт товаров' }}</div>
  </div>
  <nav class="nav">
    <a class="nav-item active" data-page="dashboard" onclick="showPage('dashboard')">
      <span class="icon">📊</span> <span>Дашборд</span>
    </a>
    <a class="nav-item" data-page="income" onclick="showPage('income')">
      <span class="icon">📥</span> <span>Приход</span>
    </a>
    <a class="nav-item" data-page="expense" onclick="showPage('expense')">
      <span class="icon">📤</span> <span>Расход</span>
    </a>
    <a class="nav-item" data-page="stock" onclick="showPage('stock')">
      <span class="icon">📦</span> <span>Остатки</span>
    </a>
    <a class="nav-item" data-page="history" onclick="showPage('history')">
      <span class="icon">📋</span> <span>История</span>
    </a>
    <a class="nav-item" data-page="products" onclick="showPage('products')">
      <span class="icon">📁</span> <span>Товары</span>
    </a>
    {% if role == 'admin' %}
    <a class="nav-item" data-page="users" onclick="showPage('users')">
      <span class="icon">👥</span> <span>Пользователи</span>
    </a>
    {% endif %}
  </nav>
  <div style="padding:16px 18px;border-top:1px solid var(--border);">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
      <div style="width:30px;height:30px;border-radius:50%;background:linear-gradient(135deg,var(--accent),var(--purple));display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;">
        {{ username[0]|upper }}
      </div>
      <div>
        <div style="font-size:12px;font-weight:600;">{{ username }}</div>
        <div style="font-size:10px;color:var(--text3);">{{ 'Администратор' if role == 'admin' else 'Продавец' }}</div>
      </div>
    </div>
    <a href="/logout" class="nav-item" style="padding:8px 0;border:none;color:var(--red);">
      <span class="icon">🚪</span> <span>Выйти</span>
    </a>
  </div>
</aside>

<!-- MAIN -->
<main class="main">
  <div class="topbar">
    <button class="burger" id="burger" onclick="toggleSidebar()" aria-label="Меню">☰</button>
    <div class="topbar-title" id="pageTitle">Дашборд</div>
  </div>

  <!-- DASHBOARD -->
  <div class="page active" id="page-dashboard">
    <div class="stats-grid">
      <div class="stat-card accent">
        <div class="stat-icon">📦</div>
        <div class="stat-label">Товаров</div>
        <div class="stat-value">{{ summary.total_products }}</div>
        <div class="stat-sub">наименований</div>
      </div>
      <div class="stat-card green">
        <div class="stat-icon">📥</div>
        <div class="stat-label">Сумма прихода</div>
        <div class="stat-value">{{ "{:,.0f}".format(summary.total_in).replace(",", " ") }}</div>
        <div class="stat-sub">сом</div>
      </div>
      <div class="stat-card red">
        <div class="stat-icon">📤</div>
        <div class="stat-label">Сумма расхода</div>
        <div class="stat-value">{{ "{:,.0f}".format(summary.total_out).replace(",", " ") }}</div>
        <div class="stat-sub">сом</div>
      </div>
      <div class="stat-card yellow">
        <div class="stat-icon">💰</div>
        <div class="stat-label">Прибыль</div>
        <div class="stat-value {{ 'profit-pos' if summary.profit >= 0 else 'profit-neg' }}">{{ "{:,.0f}".format(summary.profit).replace(",", " ") }}</div>
        <div class="stat-sub">сом</div>
      </div>
    </div>

    <!-- Quick stock table -->
    <div class="card">
      <div class="card-header">
        <div class="card-title">📦 Остатки товаров</div>
      </div>
      <table class="tbl">
        <thead><tr>
          <th>Товар</th><th class="num">Остаток</th>
          <th class="num">Ср. цена прихода</th><th class="num">Ср. цена расхода</th>
          <th class="num">Прибыль</th>
        </tr></thead>
        <tbody>
          {% for item in inventory %}
          <tr>
            <td>{{ item.name }}</td>
            <td class="num">{{ "%.2f"|format(item.stock) }}</td>
            <td class="num">{{ "%.2f"|format(item.avg_in_price) if item.avg_in_price else "—" }}</td>
            <td class="num">{{ "%.2f"|format(item.avg_out_price) if item.avg_out_price else "—" }}</td>
            <td class="num {{ 'profit-pos' if item.profit >= 0 else 'profit-neg' }}">{{ "%.2f"|format(item.profit) }}</td>
          </tr>
          {% endfor %}
          {% if not inventory %}
          <tr><td colspan="5" style="text-align:center;color:var(--text3);padding:32px;">Нет товаров</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- INCOME PAGE -->
  <div class="page" id="page-income">
    <div class="page-head">
      <div style="font-size:12px;color:var(--text3);">Приход товаров</div>
      <div class="page-head-actions">
        <button class="btn btn-secondary" onclick="openModal('uploadModal')">📎 Загрузить из файла</button>
        <button class="btn btn-primary" onclick="openModal('incomeModal')">+ Новый приход</button>
      </div>
    </div>
    <div class="card">
      <table class="tbl">
        <thead><tr>
          <th>Дата</th><th>Товар</th><th class="num">Кол-во</th>
          <th class="num">Цена</th><th class="num">Сумма</th><th></th>
        </tr></thead>
        <tbody>
          {% for t in transactions if t.type == 'in' %}
          <tr>
            <td>{{ t.date }}</td>
            <td>{{ t.product_name }}</td>
            <td class="num">{{ "%.2f"|format(t.quantity) }}</td>
            <td class="num">{{ "%.2f"|format(t.price) }}</td>
            <td class="num">{{ "%.2f"|format(t.quantity * t.price) }}</td>
            <td>
              <form method="POST" action="/delete_transaction/{{ t.id }}" onsubmit="return confirm('Удалить?')">
                <button type="submit" class="btn btn-danger btn-sm">✕</button>
              </form>
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- EXPENSE PAGE -->
  <div class="page" id="page-expense">
    <div class="page-head">
      <div style="font-size:12px;color:var(--text3);">Расход товаров</div>
      <button class="btn btn-primary" onclick="openModal('expenseModal')">+ Новый расход</button>
    </div>
    <div class="card">
      <table class="tbl">
        <thead><tr>
          <th>Дата</th><th>Товар</th><th class="num">Кол-во</th>
          <th class="num">Цена</th><th class="num">Сумма</th><th></th>
        </tr></thead>
        <tbody>
          {% for t in transactions if t.type == 'out' %}
          <tr>
            <td>{{ t.date }}</td>
            <td>{{ t.product_name }}</td>
            <td class="num">{{ "%.2f"|format(t.quantity) }}</td>
            <td class="num">{{ "%.2f"|format(t.price) }}</td>
            <td class="num">{{ "%.2f"|format(t.quantity * t.price) }}</td>
            <td>
              <form method="POST" action="/delete_transaction/{{ t.id }}" onsubmit="return confirm('Удалить?')">
                <button type="submit" class="btn btn-danger btn-sm">✕</button>
              </form>
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- STOCK PAGE -->
  <div class="page" id="page-stock">
    <div style="margin-bottom:20px;">
      <div style="font-size:12px;color:var(--text3);">Текущие остатки на складе</div>
    </div>
    <div class="card">
      <table class="tbl">
        <thead><tr>
          <th>Товар</th><th class="num">Остаток</th>
          <th class="num">Ср. цена прихода</th><th class="num">Ср. цена расхода</th>
          <th class="num">Сумма прихода</th><th class="num">Сумма расхода</th>
          <th class="num">Прибыль</th>
        </tr></thead>
        <tbody>
          {% for item in inventory %}
          <tr>
            <td>{{ item.name }}</td>
            <td class="num">{{ "%.2f"|format(item.stock) }}</td>
            <td class="num">{{ "%.2f"|format(item.avg_in_price) if item.avg_in_price else "—" }}</td>
            <td class="num">{{ "%.2f"|format(item.avg_out_price) if item.avg_out_price else "—" }}</td>
            <td class="num">{{ "%.2f"|format(item.total_in) }}</td>
            <td class="num">{{ "%.2f"|format(item.total_out) }}</td>
            <td class="num {{ 'profit-pos' if item.profit >= 0 else 'profit-neg' }}">{{ "%.2f"|format(item.profit) }}</td>
          </tr>
          {% endfor %}
          {% if not inventory %}
          <tr><td colspan="7" style="text-align:center;color:var(--text3);padding:32px;">Нет данных</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- HISTORY PAGE -->
  <div class="page" id="page-history">
    <div style="margin-bottom:20px;">
      <div style="font-size:12px;color:var(--text3);">Все операции (последние 100)</div>
    </div>
    <div class="card">
      <table class="tbl">
        <thead><tr>
          <th>Дата</th><th>Товар</th><th>Тип</th>
          <th class="num">Кол-во</th><th class="num">Цена</th><th class="num">Сумма</th><th></th>
        </tr></thead>
        <tbody>
          {% for t in transactions %}
          <tr>
            <td>{{ t.date }}</td>
            <td>{{ t.product_name }}</td>
            <td><span class="badge {{ 'badge-in' if t.type == 'in' else 'badge-out' }}">{{ 'Приход' if t.type == 'in' else 'Расход' }}</span></td>
            <td class="num">{{ "%.2f"|format(t.quantity) }}</td>
            <td class="num">{{ "%.2f"|format(t.price) }}</td>
            <td class="num">{{ "%.2f"|format(t.quantity * t.price) }}</td>
            <td>
              <form method="POST" action="/delete_transaction/{{ t.id }}" onsubmit="return confirm('Удалить?')">
                <button type="submit" class="btn btn-danger btn-sm">✕</button>
              </form>
            </td>
          </tr>
          {% endfor %}
          {% if not transactions %}
          <tr><td colspan="7" style="text-align:center;color:var(--text3);padding:32px;">Нет операций</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- PRODUCTS PAGE -->
  <div class="page" id="page-products">
    <div class="page-head">
      <div style="font-size:12px;color:var(--text3);">Справочник товаров</div>
      <button class="btn btn-primary" onclick="openModal('productModal')">+ Добавить товар</button>
    </div>
    <div class="card">
      <table class="tbl">
        <thead><tr><th>Название</th><th>Действия</th></tr></thead>
        <tbody>
          {% for p in products %}
          <tr>
            <td>{{ p.name }}</td>
            <td>
              <form method="POST" action="/delete_product/{{ p.id }}" onsubmit="return confirm('Удалить товар и все его операции?')">
                <button type="submit" class="btn btn-danger btn-sm">✕ Удалить</button>
              </form>
            </td>
          </tr>
          {% endfor %}
          {% if not products %}
          <tr><td colspan="2" style="text-align:center;color:var(--text3);padding:32px;">Нет товаров</td></tr>
          {% endif %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- USERS PAGE (admin only) -->
  {% if role == 'admin' %}
  <div class="page" id="page-users">
    <div class="page-head">
      <div style="font-size:12px;color:var(--text3);">Управление пользователями</div>
      <button class="btn btn-primary" onclick="openModal('userModal')">+ Добавить пользователя</button>
    </div>
    <div class="card">
      <table class="tbl">
        <thead><tr><th>Логин</th><th>Роль</th><th>Действия</th></tr></thead>
        <tbody>
          {% for u in users %}
          <tr>
            <td>{{ u.username }}</td>
            <td><span class="badge {{ 'badge-in' if u.role == 'admin' else 'badge-out' }}">{{ 'Админ' if u.role == 'admin' else 'Продавец' }}</span></td>
            <td style="display:flex;gap:6px;">
              <button class="btn btn-secondary btn-sm" onclick="openEditUser({{ u.id }}, '{{ u.username }}', '{{ u.role }}')">Изменить</button>
              {% if u.id != session_user_id %}
              <form method="POST" action="/delete_user/{{ u.id }}" onsubmit="return confirm('Удалить пользователя?')">
                <button type="submit" class="btn btn-danger btn-sm">✕</button>
              </form>
              {% endif %}
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endif %}
</main>

<!-- INCOME MODAL -->
<div class="modal-overlay" id="incomeModal">
  <div class="modal">
    <div class="modal-header">
      <div class="modal-title">📥 Новый приход</div>
      <span class="modal-close" onclick="closeModal('incomeModal')">✕</span>
    </div>
    <form method="POST" action="/add_transaction">
      <input type="hidden" name="type" value="in">
      <div class="modal-body">
        <div class="form-group">
          <label class="form-label">Товар</label>
          <select class="form-control" name="product_id" required>
            <option value="">Выберите...</option>
            {% for p in products %}
            <option value="{{ p.id }}">{{ p.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-row">
          <div class="form-group">
            <label class="form-label">Количество</label>
            <input class="form-control" type="number" name="quantity" step="0.01" min="0.01" required>
          </div>
          <div class="form-group">
            <label class="form-label">Цена (за ед.)</label>
            <input class="form-control" type="number" name="price" step="0.01" min="0" required>
          </div>
        </div>
        <div class="form-group">
          <label class="form-label">Дата</label>
          <input class="form-control" type="date" name="date" value="{{ today }}">
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-secondary" onclick="closeModal('incomeModal')">Отмена</button>
        <button type="submit" class="btn btn-success">✔ Записать приход</button>
      </div>
    </form>
  </div>
</div>

<!-- EXPENSE MODAL -->
<div class="modal-overlay" id="expenseModal">
  <div class="modal">
    <div class="modal-header">
      <div class="modal-title">📤 Новый расход</div>
      <span class="modal-close" onclick="closeModal('expenseModal')">✕</span>
    </div>
    <form method="POST" action="/add_transaction">
      <input type="hidden" name="type" value="out">
      <div class="modal-body">
        <div class="form-group">
          <label class="form-label">Товар</label>
          <select class="form-control" name="product_id" required>
            <option value="">Выберите...</option>
            {% for p in products %}
            <option value="{{ p.id }}">{{ p.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-row">
          <div class="form-group">
            <label class="form-label">Количество</label>
            <input class="form-control" type="number" name="quantity" step="0.01" min="0.01" required>
          </div>
          <div class="form-group">
            <label class="form-label">Цена (за ед.)</label>
            <input class="form-control" type="number" name="price" step="0.01" min="0" required>
          </div>
        </div>
        <div class="form-group">
          <label class="form-label">Дата</label>
          <input class="form-control" type="date" name="date" value="{{ today }}">
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-secondary" onclick="closeModal('expenseModal')">Отмена</button>
        <button type="submit" class="btn btn-success">✔ Записать расход</button>
      </div>
    </form>
  </div>
</div>

<!-- PRODUCT MODAL -->
<div class="modal-overlay" id="productModal">
  <div class="modal" style="width:400px;">
    <div class="modal-header">
      <div class="modal-title">📁 Новый товар</div>
      <span class="modal-close" onclick="closeModal('productModal')">✕</span>
    </div>
    <form method="POST" action="/add_product">
      <div class="modal-body">
        <div class="form-group">
          <label class="form-label">Название товара</label>
          <input class="form-control" type="text" name="name" placeholder="Введите название" required>
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-secondary" onclick="closeModal('productModal')">Отмена</button>
        <button type="submit" class="btn btn-primary">+ Добавить</button>
      </div>
    </form>
  </div>
</div>

<!-- UPLOAD MODAL -->
<div class="modal-overlay" id="uploadModal">
  <div class="modal" style="width:700px;">
    <div class="modal-header">
      <div class="modal-title">📎 Загрузить приход из файла</div>
      <span class="modal-close" onclick="closeModal('uploadModal')">✕</span>
    </div>
    <div class="modal-body">
      <div style="margin-bottom:16px;">
        <div style="font-size:12px;color:var(--text2);margin-bottom:8px;">Поддерживаемые форматы: Excel (.xlsx, .xls) и PDF</div>
        <div style="font-size:11px;color:var(--text3);margin-bottom:12px;">
          Файл должен содержать колонки: <strong>Название товара</strong>, <strong>Количество</strong>, <strong>Цена</strong>.
          Система попробует автоматически найти эти колонки.
        </div>
        <form id="uploadForm" enctype="multipart/form-data">
          <div style="display:flex;gap:10px;align-items:end;">
            <div style="flex:1;">
              <label class="form-label">Файл</label>
              <input class="form-control" type="file" name="file" id="uploadFile" accept=".xlsx,.xls,.pdf" required>
            </div>
            <div>
              <label class="form-label">Дата прихода</label>
              <input class="form-control" type="date" name="date" id="uploadDate" value="{{ today }}" style="width:160px;">
            </div>
            <button type="button" class="btn btn-primary" onclick="uploadFile()" style="white-space:nowrap;">Загрузить</button>
          </div>
        </form>
      </div>
      <div id="uploadPreview" style="display:none;">
        <div style="font-size:12px;font-weight:700;margin-bottom:8px;">Предпросмотр:</div>
        <div id="uploadAlert" style="display:none;" class="alert-info" style="padding:8px 12px;border-radius:6px;font-size:12px;margin-bottom:10px;background:rgba(79,142,247,0.1);border:1px solid rgba(79,142,247,0.3);color:var(--accent);"></div>
        <div style="max-height:300px;overflow-y:auto;">
          <table class="tbl" id="uploadTable">
            <thead><tr><th>Товар</th><th class="num">Кол-во</th><th class="num">Цена</th><th class="num">Сумма</th><th>Статус</th></tr></thead>
            <tbody id="uploadBody"></tbody>
          </table>
        </div>
        <div style="margin-top:12px;display:flex;justify-content:space-between;align-items:center;">
          <div style="font-size:13px;font-weight:700;">Итого: <span id="uploadTotal" style="color:var(--green);">0</span> сом</div>
          <button class="btn btn-success" onclick="confirmUpload()">✔ Провести приход</button>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- USER MODAL (add) -->
{% if role == 'admin' %}
<div class="modal-overlay" id="userModal">
  <div class="modal" style="width:420px;">
    <div class="modal-header">
      <div class="modal-title">👤 Новый пользователь</div>
      <span class="modal-close" onclick="closeModal('userModal')">✕</span>
    </div>
    <form method="POST" action="/add_user">
      <div class="modal-body">
        <div class="form-group">
          <label class="form-label">Логин</label>
          <input class="form-control" type="text" name="username" placeholder="Введите логин" required>
        </div>
        <div class="form-group">
          <label class="form-label">Пароль</label>
          <input class="form-control" type="password" name="password" placeholder="Введите пароль" required>
        </div>
        <div class="form-group">
          <label class="form-label">Роль</label>
          <select class="form-control" name="role">
            <option value="seller">Продавец</option>
            <option value="admin">Администратор</option>
          </select>
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-secondary" onclick="closeModal('userModal')">Отмена</button>
        <button type="submit" class="btn btn-primary">+ Добавить</button>
      </div>
    </form>
  </div>
</div>

<!-- USER EDIT MODAL -->
<div class="modal-overlay" id="editUserModal">
  <div class="modal" style="width:420px;">
    <div class="modal-header">
      <div class="modal-title">✏️ Редактировать пользователя</div>
      <span class="modal-close" onclick="closeModal('editUserModal')">✕</span>
    </div>
    <form method="POST" action="/edit_user">
      <input type="hidden" name="user_id" id="editUserId">
      <div class="modal-body">
        <div class="form-group">
          <label class="form-label">Логин</label>
          <input class="form-control" type="text" name="username" id="editUserName" required>
        </div>
        <div class="form-group">
          <label class="form-label">Новый пароль (оставьте пустым чтобы не менять)</label>
          <input class="form-control" type="password" name="password" placeholder="Новый пароль">
        </div>
        <div class="form-group">
          <label class="form-label">Роль</label>
          <select class="form-control" name="role" id="editUserRole">
            <option value="seller">Продавец</option>
            <option value="admin">Администратор</option>
          </select>
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-secondary" onclick="closeModal('editUserModal')">Отмена</button>
        <button type="submit" class="btn btn-success">✔ Сохранить</button>
      </div>
    </form>
  </div>
</div>
{% endif %}

<!-- NOTIFICATION -->
<div class="notif" id="notif"></div>

<script>
const pageTitles = {
  dashboard: 'Дашборд',
  income: 'Приход',
  expense: 'Расход',
  stock: 'Остатки',
  history: 'История операций',
  products: 'Товары',
  users: 'Пользователи'
};

function showPage(name) {
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
  document.getElementById('page-' + name).classList.add('active');
  document.querySelector('[data-page="' + name + '"]').classList.add('active');
  document.getElementById('pageTitle').textContent = pageTitles[name] || name;
  closeSidebar();
  window.scrollTo(0, 0);
}

// --- Мобильное меню ---
function toggleSidebar() {
  const open = document.getElementById('sidebar').classList.toggle('open');
  document.getElementById('sidebarBackdrop').classList.toggle('open', open);
  document.body.style.overflow = open ? 'hidden' : '';
}

function closeSidebar() {
  document.getElementById('sidebar').classList.remove('open');
  document.getElementById('sidebarBackdrop').classList.remove('open');
  document.body.style.overflow = '';
}

// При повороте экрана / переходе на десктоп меню не должно залипать открытым
window.addEventListener('resize', () => {
  if (window.innerWidth > 768) closeSidebar();
});

function openModal(id) {
  document.getElementById(id).classList.add('open');
}

function closeModal(id) {
  document.getElementById(id).classList.remove('open');
}

function openEditUser(id, username, role) {
  document.getElementById('editUserId').value = id;
  document.getElementById('editUserName').value = username;
  document.getElementById('editUserRole').value = role;
  openModal('editUserModal');
}

// Close modal on overlay click
document.querySelectorAll('.modal-overlay').forEach(overlay => {
  overlay.addEventListener('click', function(e) {
    if (e.target === this) this.classList.remove('open');
  });
});

// Close modal on Escape
document.addEventListener('keydown', function(e) {
  if (e.key === 'Escape') {
    document.querySelectorAll('.modal-overlay.open').forEach(m => m.classList.remove('open'));
    closeSidebar();
  }
});

// File upload
let uploadedRows = [];

function uploadFile() {
  const fileInput = document.getElementById('uploadFile');
  if (!fileInput.files.length) return;
  const formData = new FormData();
  formData.append('file', fileInput.files[0]);
  fetch('/parse_file', { method: 'POST', body: formData })
    .then(r => r.json())
    .then(data => {
      if (data.error) {
        alert(data.error);
        return;
      }
      uploadedRows = data.rows;
      const tbody = document.getElementById('uploadBody');
      const alertDiv = document.getElementById('uploadAlert');
      let total = 0;
      let newCount = 0;
      tbody.innerHTML = '';
      data.rows.forEach((row, i) => {
        const sum = row.quantity * row.price;
        total += sum;
        const isNew = !row.product_id;
        if (isNew) newCount++;
        tbody.innerHTML += '<tr>' +
          '<td>' + row.name + (isNew ? ' <span style="color:var(--yellow);font-size:10px;">новый</span>' : '') + '</td>' +
          '<td class="num">' + row.quantity.toFixed(2) + '</td>' +
          '<td class="num">' + row.price.toFixed(2) + '</td>' +
          '<td class="num">' + sum.toFixed(2) + '</td>' +
          '<td>' + (isNew ? '<span class="badge" style="background:rgba(247,196,79,0.15);color:var(--yellow);">Будет создан</span>' : '<span class="badge badge-in">Найден</span>') + '</td>' +
          '</tr>';
      });
      document.getElementById('uploadTotal').textContent = total.toFixed(0);
      if (newCount > 0) {
        alertDiv.style.display = 'block';
        alertDiv.textContent = 'Новых товаров: ' + newCount + ' — будут автоматически добавлены';
      } else {
        alertDiv.style.display = 'none';
      }
      document.getElementById('uploadPreview').style.display = 'block';
    })
    .catch(err => alert('Ошибка: ' + err));
}

function confirmUpload() {
  const date = document.getElementById('uploadDate').value;
  fetch('/import_income', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ rows: uploadedRows, date: date })
  })
  .then(r => r.json())
  .then(data => {
    if (data.ok) {
      window.location.reload();
    } else {
      alert(data.error || 'Ошибка');
    }
  });
}
</script>
</body>
</html>
'''

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session and 'company_id' in session:
        return redirect(url_for('index'))
    error = None
    company_code = ''
    if request.method == 'POST':
        company_code = request.form.get('company', '').strip()
        username = request.form['username'].strip()
        password = request.form['password']

        company = find_company_by_code(company_code)
        if company is None:
            # Не уточняем, что именно неверно: иначе форму можно использовать
            # для перебора существующих кодов компаний.
            error = 'Неверный код компании, логин или пароль'
        else:
            try:
                conn = connect(company['db_url'])
                user = db_fetchone(conn, 'SELECT * FROM users WHERE username = ?', (username,))
                conn.close()
            except Exception as e:
                print(f"[EasyStock] Не удалось подключиться к базе компании "
                      f"{company['code']}: {e}")
                user = None
                error = 'База компании недоступна, обратитесь к администратору'

            if error is None:
                if user and check_password_hash(user['password'], password):
                    session.clear()
                    session['company_id'] = company['id']
                    session['company_code'] = company['code']
                    session['company_name'] = company['name']
                    session['user_id'] = user['id']
                    session['username'] = user['username']
                    session['role'] = user['role']
                    return redirect(url_for('index'))
                error = 'Неверный код компании, логин или пароль'
    return render_template_string(LOGIN_TEMPLATE, error=error, is_local=not USE_PG,
                                  company_code=company_code)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    conn = get_db()
    products = db_fetchall(conn, 'SELECT * FROM products ORDER BY name')

    from datetime import date
    today = date.today().isoformat()

    inventory = db_fetchall(conn, '''
        SELECT
            p.name,
            COALESCE(SUM(CASE WHEN t.type='in' THEN t.quantity ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN t.type='out' THEN t.quantity ELSE 0 END), 0) AS stock,
            CASE WHEN SUM(CASE WHEN t.type='in' THEN t.quantity ELSE 0 END) > 0
                 THEN SUM(CASE WHEN t.type='in' THEN t.quantity * t.price ELSE 0 END)
                      / SUM(CASE WHEN t.type='in' THEN t.quantity ELSE 0 END)
                 ELSE NULL END AS avg_in_price,
            CASE WHEN SUM(CASE WHEN t.type='out' THEN t.quantity ELSE 0 END) > 0
                 THEN SUM(CASE WHEN t.type='out' THEN t.quantity * t.price ELSE 0 END)
                      / SUM(CASE WHEN t.type='out' THEN t.quantity ELSE 0 END)
                 ELSE NULL END AS avg_out_price,
            COALESCE(SUM(CASE WHEN t.type='in' THEN t.quantity * t.price ELSE 0 END), 0) AS total_in,
            COALESCE(SUM(CASE WHEN t.type='out' THEN t.quantity * t.price ELSE 0 END), 0) AS total_out,
            COALESCE(SUM(CASE WHEN t.type='out' THEN t.quantity * t.price ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN t.type='in' THEN t.quantity * t.price ELSE 0 END), 0) AS profit
        FROM products p
        LEFT JOIN transactions t ON p.id = t.product_id
        GROUP BY p.id, p.name
        ORDER BY p.name
    ''')

    row = db_fetchone(conn, '''
        SELECT
            COALESCE(SUM(CASE WHEN type='in' THEN quantity * price ELSE 0 END), 0) AS total_in,
            COALESCE(SUM(CASE WHEN type='out' THEN quantity * price ELSE 0 END), 0) AS total_out
        FROM transactions
    ''')
    summary = {
        'total_products': len(products),
        'total_in': row['total_in'],
        'total_out': row['total_out'],
        'profit': row['total_out'] - row['total_in']
    }

    transactions = db_fetchall(conn, '''
        SELECT t.*, p.name AS product_name
        FROM transactions t
        JOIN products p ON t.product_id = p.id
        ORDER BY t.date DESC, t.id DESC
        LIMIT 100
    ''')

    users = []
    if session.get('role') == 'admin':
        users = db_fetchall(conn, 'SELECT * FROM users ORDER BY id')

    conn.close()
    return render_template_string(TEMPLATE, products=products, inventory=inventory,
                                  transactions=transactions, summary=summary, today=today,
                                  role=session.get('role'), username=session.get('username'),
                                  users=users, session_user_id=session.get('user_id'),
                                  company_name=session.get('company_name'),
                                  company_code=session.get('company_code'))

@app.route('/add_product', methods=['POST'])
@login_required
def add_product():
    name = request.form['name'].strip()
    if name:
        conn = get_db()
        try:
            db_execute(conn, 'INSERT INTO products (name) VALUES (?)', (name,))
            conn.commit()
        except Exception:
            if USE_PG: conn.rollback()
        conn.close()
    return redirect(url_for('index'))

@app.route('/add_transaction', methods=['POST'])
@login_required
def add_transaction():
    product_id = request.form['product_id']
    tx_type = request.form['type']
    quantity = float(request.form['quantity'])
    price = float(request.form['price'])
    date = request.form.get('date', '')
    conn = get_db()
    if date:
        db_execute(conn, 'INSERT INTO transactions (product_id, type, quantity, price, date) VALUES (?, ?, ?, ?, ?)',
                   (product_id, tx_type, quantity, price, date))
    else:
        db_execute(conn, 'INSERT INTO transactions (product_id, type, quantity, price) VALUES (?, ?, ?, ?)',
                   (product_id, tx_type, quantity, price))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/delete_transaction/<int:tx_id>', methods=['POST'])
@login_required
def delete_transaction(tx_id):
    conn = get_db()
    db_execute(conn, 'DELETE FROM transactions WHERE id = ?', (tx_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/delete_product/<int:product_id>', methods=['POST'])
@login_required
def delete_product(product_id):
    conn = get_db()
    db_execute(conn, 'DELETE FROM transactions WHERE product_id = ?', (product_id,))
    db_execute(conn, 'DELETE FROM products WHERE id = ?', (product_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/add_user', methods=['POST'])
@admin_required
def add_user():
    username = request.form['username'].strip()
    password = request.form['password']
    role = request.form['role']
    if username and password:
        conn = get_db()
        try:
            db_execute(conn, 'INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                       (username, generate_password_hash(password), role))
            conn.commit()
        except Exception:
            if USE_PG: conn.rollback()
        conn.close()
    return redirect(url_for('index'))

@app.route('/edit_user', methods=['POST'])
@admin_required
def edit_user():
    user_id = int(request.form['user_id'])
    username = request.form['username'].strip()
    password = request.form.get('password', '')
    role = request.form['role']
    conn = get_db()
    if password:
        db_execute(conn, 'UPDATE users SET username=?, password=?, role=? WHERE id=?',
                   (username, generate_password_hash(password), role, user_id))
    else:
        db_execute(conn, 'UPDATE users SET username=?, role=? WHERE id=?',
                   (username, role, user_id))
    conn.commit()
    conn.close()
    if user_id == session.get('user_id'):
        session['username'] = username
        session['role'] = role
    return redirect(url_for('index'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    if user_id != session.get('user_id'):
        conn = get_db()
        db_execute(conn, 'DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
    return redirect(url_for('index'))

@app.route('/parse_file', methods=['POST'])
@login_required
def parse_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не выбран'})
    f = request.files['file']
    fname = f.filename.lower()

    try:
        if fname.endswith(('.xlsx', '.xls')):
            rows = parse_excel(f)
        elif fname.endswith('.pdf'):
            rows = parse_pdf(f)
        else:
            return jsonify({'error': 'Неподдерживаемый формат. Используйте .xlsx или .pdf'})
    except Exception as e:
        return jsonify({'error': f'Ошибка чтения файла: {str(e)}'})

    if not rows:
        return jsonify({'error': 'Не удалось найти данные в файле. Убедитесь что есть колонки: название, количество, цена.'})

    # Match with existing products
    conn = get_db()
    products = db_fetchall(conn, 'SELECT id, name FROM products')
    conn.close()
    product_map = {p['name'].lower().strip(): p['id'] for p in products}

    for row in rows:
        name_lower = row['name'].lower().strip()
        row['product_id'] = product_map.get(name_lower)

    return jsonify({'rows': rows})


def parse_excel(f):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    f.save(tmp.name)
    tmp.close()

    wb = openpyxl.load_workbook(tmp.name, read_only=True)
    ws = wb.active
    rows_data = []

    # Find header row
    header_row = None
    name_col = qty_col = price_col = None
    name_keywords = ['название', 'наименование', 'товар', 'продукт', 'name', 'item', 'позиция']
    qty_keywords = ['количество', 'кол-во', 'кол', 'qty', 'quantity', 'шт']
    price_keywords = ['цена', 'стоимость', 'price', 'cost', 'сумма за ед']

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        for j, cell in enumerate(row):
            val = str(cell.value or '').lower().strip()
            if any(k in val for k in name_keywords):
                name_col = j
                header_row = i
            if any(k in val for k in qty_keywords):
                qty_col = j
            if any(k in val for k in price_keywords):
                price_col = j
        if header_row:
            break

    # If no header found, assume columns: A=name, B=qty, C=price
    if header_row is None:
        header_row = 0
        name_col = 0
        qty_col = 1
        price_col = 2

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        cells = list(row)
        if name_col is not None and name_col < len(cells):
            name = str(cells[name_col].value or '').strip()
        else:
            continue
        if not name or name.lower() in ('итого', 'total', 'всего', ''):
            continue

        qty = 0
        price = 0
        if qty_col is not None and qty_col < len(cells):
            try:
                qty = float(cells[qty_col].value or 0)
            except (ValueError, TypeError):
                continue
        if price_col is not None and price_col < len(cells):
            try:
                price = float(cells[price_col].value or 0)
            except (ValueError, TypeError):
                price = 0

        if qty > 0 and name:
            rows_data.append({'name': name, 'quantity': qty, 'price': price})

    wb.close()
    os.unlink(tmp.name)
    return rows_data


def parse_pdf(f):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    f.save(tmp.name)
    tmp.close()

    rows_data = []
    with pdfplumber.open(tmp.name) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                # Find header
                header_idx = None
                name_col = qty_col = price_col = None
                name_keywords = ['название', 'наименование', 'товар', 'продукт', 'позиция']
                qty_keywords = ['количество', 'кол-во', 'кол', 'шт']
                price_keywords = ['цена', 'стоимость', 'сумма за ед']

                for i, row in enumerate(table):
                    for j, cell in enumerate(row):
                        val = str(cell or '').lower().strip()
                        if any(k in val for k in name_keywords):
                            name_col = j
                            header_idx = i
                        if any(k in val for k in qty_keywords):
                            qty_col = j
                        if any(k in val for k in price_keywords):
                            price_col = j
                    if header_idx is not None:
                        break

                if header_idx is None:
                    # Try: first col = name, find numbers in other cols
                    header_idx = 0
                    name_col = 0
                    qty_col = 1 if len(table[0]) > 1 else None
                    price_col = 2 if len(table[0]) > 2 else None

                for row in table[header_idx + 1:]:
                    if name_col is not None and name_col < len(row):
                        name = str(row[name_col] or '').strip()
                    else:
                        continue
                    if not name or name.lower() in ('итого', 'total', 'всего'):
                        continue

                    qty = 0
                    price = 0
                    if qty_col is not None and qty_col < len(row):
                        try:
                            qty = float(re.sub(r'[^\d.,]', '', str(row[qty_col] or '0')).replace(',', '.') or 0)
                        except (ValueError, TypeError):
                            continue
                    if price_col is not None and price_col < len(row):
                        try:
                            price = float(re.sub(r'[^\d.,]', '', str(row[price_col] or '0')).replace(',', '.') or 0)
                        except (ValueError, TypeError):
                            price = 0

                    if qty > 0 and name:
                        rows_data.append({'name': name, 'quantity': qty, 'price': price})

    os.unlink(tmp.name)
    return rows_data


@app.route('/import_income', methods=['POST'])
@login_required
def import_income():
    data = request.get_json()
    rows = data.get('rows', [])
    date = data.get('date', '')
    conn = get_db()

    for row in rows:
        name = row['name'].strip()
        qty = float(row['quantity'])
        price = float(row['price'])
        product_id = row.get('product_id')

        # Create product if not exists
        if not product_id:
            try:
                db_execute(conn, 'INSERT INTO products (name) VALUES (?)', (name,))
                conn.commit()
            except Exception:
                if USE_PG: conn.rollback()
            p = db_fetchone(conn, 'SELECT id FROM products WHERE name = ?', (name,))
            product_id = p['id']

        if date:
            db_execute(conn, 'INSERT INTO transactions (product_id, type, quantity, price, date) VALUES (?, ?, ?, ?, ?)',
                       (product_id, 'in', qty, price, date))
        else:
            db_execute(conn, 'INSERT INTO transactions (product_id, type, quantity, price) VALUES (?, ?, ?, ?)',
                       (product_id, 'in', qty, price))

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'count': len(rows)})


# --- Панель владельца сервиса ---------------------------------------------

SUPERADMIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EasyStock — Компании</title>
<style>
:root {
  --bg:#0e1117; --bg2:#161b27; --bg3:#1e2535; --border:#2a3347; --border2:#374056;
  --text:#e8ecf4; --text2:#8b95b0; --text3:#5a6480;
  --accent:#4f8ef7; --accent2:#3b6fd4; --green:#2ecc8a; --red:#f74f4f; --yellow:#f7c44f;
  --radius:10px; --radius2:6px;
}
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:var(--bg); color:var(--text);
       font-size:14px; min-height:100vh; min-height:100dvh; padding:24px; }
.wrap { max-width:1000px; margin:0 auto; }
.head { display:flex; justify-content:space-between; align-items:center;
        gap:12px; margin-bottom:22px; flex-wrap:wrap; }
.head h1 { font-size:19px; color:var(--accent); }
.head .sub { font-size:11px; color:var(--text3); margin-top:2px; }
.card { background:var(--bg2); border:1px solid var(--border);
        border-radius:var(--radius); margin-bottom:20px; overflow-x:auto; }
.card-h { padding:14px 18px; border-bottom:1px solid var(--border); font-size:13px; font-weight:700; }
.card-b { padding:18px; }
.tbl { width:100%; border-collapse:collapse; min-width:560px; }
.tbl th { text-align:left; padding:10px 14px; font-size:10px; font-weight:700;
          letter-spacing:.08em; text-transform:uppercase; color:var(--text3);
          background:var(--bg3); border-bottom:1px solid var(--border); }
.tbl td { padding:11px 14px; font-size:12px; border-bottom:1px solid var(--border); }
.tbl tr:last-child td { border-bottom:none; }
.badge { display:inline-flex; padding:2px 8px; border-radius:20px; font-size:10px; font-weight:700; }
.badge-on { background:rgba(46,204,138,.15); color:var(--green); }
.badge-off { background:rgba(247,79,79,.15); color:var(--red); }
.form-group { margin-bottom:14px; }
.form-label { display:block; font-size:11px; font-weight:600; color:var(--text2); margin-bottom:5px; }
.form-hint { font-size:10px; color:var(--text3); margin-top:4px; }
.form-control { width:100%; padding:10px 12px; background:var(--bg3);
                border:1px solid var(--border2); border-radius:var(--radius2);
                color:var(--text); font-family:inherit; font-size:14px; outline:none; }
.form-control:focus { border-color:var(--accent); }
.row2 { display:grid; grid-template-columns:1fr 1fr; gap:12px; }
.btn { display:inline-flex; align-items:center; gap:6px; padding:10px 16px;
       border-radius:var(--radius2); font-size:12px; font-weight:600; cursor:pointer;
       border:none; font-family:inherit; color:#fff; min-height:40px; }
.btn-primary { background:var(--accent); }
.btn-danger { background:#8a1a1a; color:var(--red); }
.btn-secondary { background:var(--bg3); color:var(--text); border:1px solid var(--border2); }
.btn-sm { padding:7px 11px; min-height:34px; font-size:11px; }
.msg { padding:10px 14px; border-radius:var(--radius2); font-size:12px; margin-bottom:16px; }
.msg-err { background:rgba(247,79,79,.1); border:1px solid rgba(247,79,79,.3); color:var(--red); }
.msg-ok { background:rgba(46,204,138,.1); border:1px solid rgba(46,204,138,.3); color:var(--green); }
.acts { display:flex; gap:6px; }
@media (max-width:768px) {
  body { padding:14px; }
  .row2 { grid-template-columns:1fr; }
  .form-control { font-size:16px; }
}
</style>
</head>
<body>
<div class="wrap">
  <div class="head">
    <div>
      <h1>EasyStock — Компании</h1>
      <div class="sub">Панель владельца сервиса</div>
    </div>
    <a href="/superadmin/logout" class="btn btn-secondary">Выйти</a>
  </div>

  {% if error %}<div class="msg msg-err">{{ error }}</div>{% endif %}
  {% if ok %}<div class="msg msg-ok">{{ ok }}</div>{% endif %}

  <div class="card">
    <div class="card-h">Компании ({{ companies|length }})</div>
    <table class="tbl">
      <thead><tr><th>Код</th><th>Название</th><th>Создана</th><th>Статус</th><th>Действия</th></tr></thead>
      <tbody>
        {% for c in companies %}
        <tr>
          <td><code>{{ c.code }}</code></td>
          <td>{{ c.name }}</td>
          <td>{{ c.created_at }}</td>
          <td>
            {% if c.is_active %}<span class="badge badge-on">Активна</span>
            {% else %}<span class="badge badge-off">Отключена</span>{% endif %}
          </td>
          <td>
            <div class="acts">
              <form method="POST" action="/superadmin/toggle/{{ c.id }}">
                <button class="btn btn-secondary btn-sm">
                  {{ 'Отключить' if c.is_active else 'Включить' }}
                </button>
              </form>
              <form method="POST" action="/superadmin/delete/{{ c.id }}"
                    onsubmit="return confirm('Убрать компанию {{ c.name }} из реестра? Её база и данные останутся нетронутыми.')">
                <button class="btn btn-danger btn-sm">Убрать</button>
              </form>
            </div>
          </td>
        </tr>
        {% endfor %}
        {% if not companies %}
        <tr><td colspan="5" style="text-align:center;color:var(--text3);padding:32px;">
          Компаний пока нет — добавьте первую ниже
        </td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>

  <div class="card">
    <div class="card-h">Добавить компанию</div>
    <div class="card-b">
      <form method="POST" action="/superadmin/add">
        <div class="row2">
          <div class="form-group">
            <label class="form-label">Код компании</label>
            <input class="form-control" name="code" placeholder="acme" required
                   autocapitalize="off" autocorrect="off">
            <div class="form-hint">Латиницей, без пробелов. Его сотрудники вводят при входе.</div>
          </div>
          <div class="form-group">
            <label class="form-label">Название</label>
            <input class="form-control" name="name" placeholder="ООО Акме" required>
          </div>
        </div>
        <div class="form-group">
          <label class="form-label">Строка подключения к базе компании</label>
          <input class="form-control" name="db_url" required
                 placeholder="{{ 'postgresql://...' if use_pg else 'company_acme.db' }}">
          <div class="form-hint">
            {% if use_pg %}
              Создайте отдельную базу в консоли Neon и вставьте её строку подключения.
              Схема развернётся автоматически.
            {% else %}
              Локально достаточно имени файла, например company_acme.db
            {% endif %}
          </div>
        </div>
        <div class="row2">
          <div class="form-group">
            <label class="form-label">Пароль администратора компании</label>
            <input class="form-control" type="text" name="admin_password" required>
            <div class="form-hint">Логин будет admin</div>
          </div>
          <div class="form-group">
            <label class="form-label">Пароль продавца</label>
            <input class="form-control" type="text" name="seller_password" required>
            <div class="form-hint">Логин будет seller</div>
          </div>
        </div>
        <button class="btn btn-primary">Создать компанию</button>
      </form>
    </div>
  </div>
</div>
</body>
</html>
'''

SUPERADMIN_LOGIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EasyStock — Вход владельца</title>
<style>
:root { --bg:#0e1117; --bg2:#161b27; --bg3:#1e2535; --border:#2a3347; --border2:#374056;
        --text:#e8ecf4; --text2:#8b95b0; --text3:#5a6480; --accent:#4f8ef7; --accent2:#3b6fd4;
        --red:#f74f4f; --radius:10px; --radius2:6px; }
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:var(--bg); color:var(--text);
       min-height:100vh; min-height:100dvh; display:flex; align-items:center; justify-content:center; padding:16px; }
.card { background:var(--bg2); border:1px solid var(--border); border-radius:var(--radius);
        padding:40px; width:380px; max-width:100%; box-shadow:0 4px 24px rgba(0,0,0,.4); }
h1 { font-size:20px; color:var(--accent); text-align:center; }
.sub { font-size:11px; color:var(--text3); text-align:center; margin:4px 0 26px; }
.form-group { margin-bottom:16px; }
.form-label { display:block; font-size:11px; font-weight:600; color:var(--text2); margin-bottom:5px; }
.form-control { width:100%; padding:11px 14px; background:var(--bg3); border:1px solid var(--border2);
                border-radius:var(--radius2); color:var(--text); font-family:inherit; font-size:14px; outline:none; }
.form-control:focus { border-color:var(--accent); }
.btn { width:100%; padding:13px; background:var(--accent); color:#fff; border:none;
       border-radius:var(--radius2); font-size:14px; font-weight:600; cursor:pointer; font-family:inherit; }
.btn:hover { background:var(--accent2); }
.err { background:rgba(247,79,79,.1); border:1px solid rgba(247,79,79,.3); color:var(--red);
       padding:9px 12px; border-radius:var(--radius2); font-size:12px; margin-bottom:16px; text-align:center; }
.back { display:block; text-align:center; margin-top:18px; font-size:11px; color:var(--text3); text-decoration:none; }
@media (max-width:480px) { .card { padding:28px 22px; } .form-control { font-size:16px; } }
</style>
</head>
<body>
<div class="card">
  <h1>EasyStock</h1>
  <div class="sub">Вход для владельца сервиса</div>
  {% if error %}<div class="err">{{ error }}</div>{% endif %}
  <form method="POST">
    <div class="form-group">
      <label class="form-label">Логин</label>
      <input class="form-control" name="username" required autofocus autocapitalize="off">
    </div>
    <div class="form-group">
      <label class="form-label">Пароль</label>
      <input class="form-control" type="password" name="password" required>
    </div>
    <button class="btn">Войти</button>
  </form>
  <a class="back" href="/login">← Вход для сотрудников компании</a>
</div>
</body>
</html>
'''


@app.route('/superadmin/login', methods=['GET', 'POST'])
def superadmin_login():
    if session.get('is_superadmin'):
        return redirect(url_for('superadmin'))
    error = None
    if request.method == 'POST':
        if (request.form.get('username', '').strip() == SUPERADMIN_USERNAME
                and request.form.get('password', '') == SUPERADMIN_PASSWORD):
            session.clear()
            session['is_superadmin'] = True
            return redirect(url_for('superadmin'))
        error = 'Неверный логин или пароль'
    return render_template_string(SUPERADMIN_LOGIN_TEMPLATE, error=error)


@app.route('/superadmin/logout')
def superadmin_logout():
    session.clear()
    return redirect(url_for('superadmin_login'))


@app.route('/superadmin')
@superadmin_required
def superadmin():
    return render_template_string(
        SUPERADMIN_TEMPLATE, companies=all_companies(), use_pg=USE_PG,
        error=request.args.get('error'), ok=request.args.get('ok'))


@app.route('/superadmin/add', methods=['POST'])
@superadmin_required
def superadmin_add():
    code = request.form.get('code', '').strip().lower()
    name = request.form.get('name', '').strip()
    db_url = request.form.get('db_url', '').strip()
    admin_password = request.form.get('admin_password', '')
    seller_password = request.form.get('seller_password', '')

    if not re.fullmatch(r'[a-z0-9][a-z0-9_-]{1,31}', code):
        return redirect(url_for('superadmin', error='Код: латиница, цифры, дефис; от 2 до 32 знаков'))

    conn = get_registry()
    if db_fetchone(conn, 'SELECT id FROM companies WHERE code = ?', (code,)):
        conn.close()
        return redirect(url_for('superadmin', error=f'Компания с кодом {code} уже есть'))
    conn.close()

    # Схему разворачиваем ДО записи в реестр: если база недоступна,
    # в реестре не останется компании, в которую нельзя войти.
    try:
        init_company_db(db_url, admin_password, seller_password)
    except Exception as e:
        return redirect(url_for('superadmin', error=f'База недоступна: {e}'))

    conn = get_registry()
    db_execute(conn, 'INSERT INTO companies (code, name, db_url) VALUES (?, ?, ?)',
               (code, name, db_url))
    conn.commit()
    conn.close()
    forget_companies()
    return redirect(url_for('superadmin', ok=f'Компания «{name}» создана, код входа: {code}'))


@app.route('/superadmin/toggle/<int:company_id>', methods=['POST'])
@superadmin_required
def superadmin_toggle(company_id):
    conn = get_registry()
    db_execute(conn,
               'UPDATE companies SET is_active = 1 - is_active WHERE id = ?',
               (company_id,))
    conn.commit()
    conn.close()
    forget_companies()
    return redirect(url_for('superadmin'))


@app.route('/superadmin/delete/<int:company_id>', methods=['POST'])
@superadmin_required
def superadmin_delete(company_id):
    # Удаляем только запись реестра. База компании и все её данные остаются:
    # снести чужую базу одним кликом — слишком опасная операция.
    conn = get_registry()
    db_execute(conn, 'DELETE FROM companies WHERE id = ?', (company_id,))
    conn.commit()
    conn.close()
    forget_companies()
    return redirect(url_for('superadmin', ok='Компания убрана из реестра, её база не тронута'))


init_registry()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
